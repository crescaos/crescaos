"""Read current dashboard structure in detail."""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook(r'c:\Users\12132\Desktop\DigiFacil\DigiFacil\Todo_Directo_Founder_Financial_Dashboard.xlsx', data_only=True)

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n{'='*60}")
    print(f"Sheet: {sname} ({ws.max_row} rows x {ws.max_column} cols)")
    print(f"{'='*60}")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=False):
        vals = []
        for c in row:
            v = c.value
            if v is not None:
                vals.append(f"{c.coordinate}={str(v)[:40]}")
        if vals:
            print("  ".join(vals))
