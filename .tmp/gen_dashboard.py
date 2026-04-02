"""Generate updated Cresca Financial Dashboard — 3-Tier Model with 20%/6% + $600 Base."""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter

OUT = r'c:\Users\12132\Desktop\DigiFacil\DigiFacil\Todo_Directo_Founder_Financial_Dashboard.xlsx'
wb = Workbook()

# Style definitions
BLUE_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
GREEN_FILL = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
LIGHT_GREEN = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
GRAY_FILL = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
DARK_FILL = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
WHITE_FONT = Font(color='FFFFFF', bold=True, size=11)
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color='1A1A2E')
INPUT_FILL = PatternFill(start_color='EBF5FF', end_color='EBF5FF', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, max_col, fill=BLUE_FILL, font=WHITE_FONT):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_data_cell(ws, row, col, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    return cell

# ============================================
# SHEET 1: README
# ============================================
ws_readme = wb.active
ws_readme.title = 'README'
ws_readme.column_dimensions['A'].width = 100

readme_lines = [
    ('Cresca — Founder Financial Dashboard', TITLE_FONT),
    ('Sistemas de Crecimiento Empresarial', Font(italic=True, size=11)),
    ('', None),
    ('How to use:', Font(bold=True, size=12)),
    ("1) Fill the blue cells in 'Clients' sheet for each new client", None),
    ("2) Maintain status as Active / Paused / Cancelled in 'Clients'", None),
    ("3) 'MRR Summary' and 'Commission Liability' update automatically", None),
    ("4) Use 'Capacity Planning' to monitor build capacity", None),
    ("5) Track MRR milestones for bonus pool in 'Targets'", None),
    ('', None),
    ('Version 3.0 — March 2026 | 3-Tier Model: STARTER / GROWTH / PRO AUTOMATION', Font(italic=True)),
    ('Changes: 3-tier pricing, founding rep comp (20%/6%/$600 base), booking as universal module', Font(italic=True)),
    ('', None),
    ('Founding Rep ES Compensation:', Font(bold=True, size=12)),
    ('Base: $600/mo | Setup: 20% | Residual: 6% | Clawback: 50% in 60 days', None),
]
for i, (text, font) in enumerate(readme_lines, 1):
    cell = ws_readme.cell(row=i, column=1, value=text)
    if font:
        cell.font = font

# ============================================
# SHEET 2: ASSUMPTIONS
# ============================================
ws_assume = wb.create_sheet('Assumptions')
ws_assume.column_dimensions['A'].width = 40
ws_assume.column_dimensions['B'].width = 20
ws_assume.column_dimensions['C'].width = 30
ws_assume.column_dimensions['D'].width = 22

ws_assume.cell(row=1, column=1, value='Cresca — Pricing Assumptions (3-Tier)').font = TITLE_FONT

# Package pricing
ws_assume.cell(row=3, column=1, value='Package Pricing').font = Font(bold=True, size=12)
headers = ['Package', 'Setup Fee ($)', 'Monthly Fee ($)', 'Workflows Included']
for j, h in enumerate(headers, 1):
    ws_assume.cell(row=4, column=j, value=h)
style_header_row(ws_assume, 4, 4)

packages = [
    ['STARTER — DIY Lite', 0, 9.99, '1 Lite'],
    ['STARTER — Starter Assist', 25, 19.99, '1 Lite'],
    ['STARTER — Starter Presence', 75, 29.99, '1 Lite'],
    ['STARTER — Starter Pro', 150, 49.99, '1 Lite'],
    ['GROWTH — Foundation', 250, 50, '1 Lite'],
    ['GROWTH — Foundation +1 WF', 350, 75, '1 Full'],
    ['GROWTH — Foundation +2 WF', 450, 100, '2 Full'],
    ['PRO AUTOMATION (avg)', 1500, 500, '4+'],
    ['Add-On: Marketing Digital', 200, 250, '—'],
]
for i, pkg in enumerate(packages, 5):
    for j, val in enumerate(pkg, 1):
        cell = ws_assume.cell(row=i, column=j, value=val)
        cell.border = THIN_BORDER
        if j >= 2 and j <= 3:
            cell.fill = INPUT_FILL
            cell.number_format = '$#,##0.00'

# Commission structure — FOUNDING REP ES
r_comm = 5 + len(packages) + 2
ws_assume.cell(row=r_comm, column=1, value='Founding Rep ES Compensation').font = Font(bold=True, size=12)
comm_headers = ['Element', 'Value', 'Notes']
for j, h in enumerate(comm_headers, 1):
    ws_assume.cell(row=r_comm+1, column=j, value=h)
style_header_row(ws_assume, r_comm+1, 3)

comp_data = [
    ['Base Salary', '$600/mo', 'Guaranteed, paid monthly'],
    ['Setup Commission', '20%', 'On collected setup fees only'],
    ['Residual Commission', '6%', 'On collected MRR, while client active'],
    ['Clawback', '50%', 'Setup commission if cancel in 60d (bad fit/oversell)'],
    ['Payout Rule', 'Collected only', 'No commission on invoiced-not-collected'],
    ['Exclusions', 'Pass-through', 'No commission on ad spend, 3rd party fees, refunds'],
]
for i, row in enumerate(comp_data, r_comm+2):
    for j, val in enumerate(row, 1):
        cell = ws_assume.cell(row=i, column=j, value=val)
        cell.border = THIN_BORDER

# Capacity
r_cap = r_comm + 2 + len(comp_data) + 2
ws_assume.cell(row=r_cap, column=1, value='Operational Capacity').font = Font(bold=True, size=12)
ws_assume.cell(row=r_cap+1, column=1, value='Max Concurrent GROWTH Builds')
ws_assume.cell(row=r_cap+1, column=2, value=5).fill = INPUT_FILL
ws_assume.cell(row=r_cap+2, column=1, value='Max New Sales/Week')
ws_assume.cell(row=r_cap+2, column=2, value=2).fill = INPUT_FILL
ws_assume.cell(row=r_cap+3, column=1, value='STARTER = high volume (minimal build)')
ws_assume.cell(row=r_cap+4, column=1, value='PRO AUTOMATION = 1-2 concurrent')

# ============================================
# SHEET 3: CLIENTS
# ============================================
ws_clients = wb.create_sheet('Clients')
client_headers = [
    'Client Name', 'Package/Tier', 'Start Date', 'Status',
    'Rep Name', 'Setup Fee ($)', 'Monthly Fee ($)',
    'Variable Cost ($/mo)', 'Setup Commission ($)', 'Residual Commission ($/mo)',
    'Effective MRR ($)', 'Gross Margin %', 'Payment Plan', 'Notes'
]
for j, h in enumerate(client_headers, 1):
    ws_clients.cell(row=1, column=j, value=h)
    ws_clients.column_dimensions[get_column_letter(j)].width = max(len(h) + 2, 14)
style_header_row(ws_clients, 1, len(client_headers))

# Data validation lists
from openpyxl.worksheet.datavalidation import DataValidation
dv_pkg = DataValidation(type='list', formula1='"STARTER DIY Lite,STARTER Assist,STARTER Presence,STARTER Pro,GROWTH Foundation,GROWTH +1 WF,GROWTH +2 WF,PRO AUTOMATION,Add-On Marketing"')
dv_status = DataValidation(type='list', formula1='"Active,Paused,Cancelled,Onboarding"')
dv_payment = DataValidation(type='list', formula1='"Monthly,6-Month Upfront,12-Month Upfront,Flex 50/50"')
ws_clients.add_data_validation(dv_pkg)
ws_clients.add_data_validation(dv_status)
ws_clients.add_data_validation(dv_payment)

for r in range(2, 102):
    dv_pkg.add(ws_clients.cell(row=r, column=2))
    dv_status.add(ws_clients.cell(row=r, column=4))
    dv_payment.add(ws_clients.cell(row=r, column=13))
    # Input cells get blue fill
    for c in [1,2,3,4,5,6,7,8,13,14]:
        ws_clients.cell(row=r, column=c).fill = INPUT_FILL
    # Setup Commission = Setup Fee * 20%
    ws_clients.cell(row=r, column=9).value = f'=IF(F{r}="","",F{r}*0.20)'
    ws_clients.cell(row=r, column=9).number_format = '$#,##0.00'
    # Residual Commission = Monthly Fee * 6%
    ws_clients.cell(row=r, column=10).value = f'=IF(G{r}="","",G{r}*0.06)'
    ws_clients.cell(row=r, column=10).number_format = '$#,##0.00'
    # Effective MRR = Monthly Fee (if Active)
    ws_clients.cell(row=r, column=11).value = f'=IF(D{r}="Active",G{r},0)'
    ws_clients.cell(row=r, column=11).number_format = '$#,##0.00'
    # Gross Margin = (Monthly - Variable - Residual) / Monthly
    ws_clients.cell(row=r, column=12).value = f'=IF(G{r}="","",IF(G{r}=0,"",(G{r}-H{r}-J{r})/G{r}))'
    ws_clients.cell(row=r, column=12).number_format = '0.0%'
    # Style formula cells
    for c in [9,10,11,12]:
        ws_clients.cell(row=r, column=c).fill = LIGHT_GREEN

# ============================================
# SHEET 4: MRR SUMMARY
# ============================================
ws_mrr = wb.create_sheet('MRR Summary')
ws_mrr.column_dimensions['A'].width = 35
ws_mrr.column_dimensions['B'].width = 20

ws_mrr.cell(row=1, column=1, value='Cresca — MRR Summary').font = TITLE_FONT

mrr_items = [
    ('Total Active Clients', '=COUNTIF(Clients!D2:D101,"Active")'),
    ('Onboarding Clients', '=COUNTIF(Clients!D2:D101,"Onboarding")'),
    ('Paused Clients', '=COUNTIF(Clients!D2:D101,"Paused")'),
    ('Cancelled Clients', '=COUNTIF(Clients!D2:D101,"Cancelled")'),
    ('', ''),
    ('MRR (Monthly Recurring Revenue)', '=SUM(Clients!K2:K101)'),
    ('ARR (Annual Recurring Revenue)', '=B8*12'),
    ('Average Revenue Per User (ARPU)', '=IF(B3=0,"",B8/B3)'),
    ('', ''),
    ('Monthly Variable Costs', '=SUMPRODUCT((Clients!D2:D101="Active")*Clients!H2:H101)'),
    ('Monthly Residual Commissions (6%)', '=SUMPRODUCT((Clients!D2:D101="Active")*Clients!J2:J101)'),
    ('Rep Base Salary', 600),
    ('Total Monthly Costs', '=B12+B13+B14'),
    ('', ''),
    ('Monthly Gross Profit', '=B8-B15'),
    ('Gross Margin %', '=IF(B8=0,"",B17/B8)'),
    ('', ''),
    ('Total Setup Fees Collected', '=SUMPRODUCT((Clients!A2:A101<>"")*Clients!F2:F101)'),
    ('Total Setup Commissions Paid (20%)', '=SUM(Clients!I2:I101)'),
    ('Net Setup Revenue', '=B20-B21'),
]
for i, (label, formula) in enumerate(mrr_items, 3):
    ws_mrr.cell(row=i, column=1, value=label)
    if label:
        ws_mrr.cell(row=i, column=1).font = Font(bold=True) if 'MRR' in label or 'Total' in label or 'Gross' in label else Font()
    if formula:
        cell = ws_mrr.cell(row=i, column=2, value=formula)
        if '%' in label:
            cell.number_format = '0.0%'
        elif isinstance(formula, str) and formula.startswith('='):
            cell.number_format = '$#,##0.00'
        elif isinstance(formula, (int, float)):
            cell.number_format = '$#,##0.00'
        cell.fill = LIGHT_GREEN

# ============================================
# SHEET 5: COMMISSION LIABILITY
# ============================================
ws_comm = wb.create_sheet('Commission Liability')
ws_comm.column_dimensions['A'].width = 35
ws_comm.column_dimensions['B'].width = 20

ws_comm.cell(row=1, column=1, value='Cresca — Commission Liability (Founding Rep ES)').font = TITLE_FONT

comm_items = [
    ('Founding Rep Compensation', '', True),
    ('Base Salary (guaranteed)', '$600/mo', False),
    ('Setup Commission Rate', '20%', False),
    ('Residual Commission Rate', '6%', False),
    ('Clawback', '50% setup if cancel in 60 days', False),
    ('Payout Rule', 'Collected revenue only', False),
    ('', '', False),
    ('Monthly Liability', '', True),
    ('Rep Base Salary', 600, False),
    ('Total Residual Payouts (6%)', '=SUM(Clients!J2:J101)', False),
    ('Total Monthly Commission Liability', '=B12+B13', False),
    ('As % of MRR', '=IF(\'MRR Summary\'!B8=0,"",B14/\'MRR Summary\'!B8)', False),
    ('', '', False),
    ('Setup Liability (YTD)', '', True),
    ('Total Setup Commissions (20%)', '=SUM(Clients!I2:I101)', False),
]
for i, (label, val, is_header) in enumerate(comm_items, 3):
    cell_a = ws_comm.cell(row=i, column=1, value=label)
    if is_header:
        cell_a.font = Font(bold=True, size=12)
    cell_b = ws_comm.cell(row=i, column=2, value=val)
    if isinstance(val, str) and val.startswith('='):
        cell_b.number_format = '$#,##0.00' if '%' not in label else '0.0%'
        cell_b.fill = LIGHT_GREEN
    elif isinstance(val, (int, float)):
        cell_b.number_format = '$#,##0.00'
        cell_b.fill = LIGHT_GREEN

# ============================================
# SHEET 6: P&L
# ============================================
ws_pl = wb.create_sheet('P&L')
ws_pl.column_dimensions['A'].width = 35
ws_pl.column_dimensions['B'].width = 20

ws_pl.cell(row=1, column=1, value='Cresca — Profit & Loss Statement').font = TITLE_FONT
ws_pl.cell(row=2, column=1, value='Monthly').font = Font(bold=True)
ws_pl.cell(row=2, column=2, value='Amount').font = Font(bold=True)

pl_items = [
    ('REVENUE', '', True),
    ('MRR (Recurring)', "='MRR Summary'!B8", False),
    ('Setup Fees (Non-Recurring)', '', False),
    ('Total Revenue', '=B5+B6', False),
    ('', '', False),
    ('COST OF SERVICE', '', True),
    ('Variable Costs', "='MRR Summary'!B12", False),
    ('Platform/Tool Subscriptions', '', False),
    ('Total COGS', '=B11+B12', False),
    ('', '', False),
    ('GROSS PROFIT', '=B8-B14', True),
    ('Gross Margin %', '=IF(B8=0,"",B16/B8)', False),
    ('', '', False),
    ('OPERATING EXPENSES', '', True),
    ('Rep Base Salary', "='Commission Liability'!B12", False),
    ('Residual Commissions (6%)', "='Commission Liability'!B13", False),
    ('Setup Commissions (amortized)', '', False),
    ('Marketing/Advertising', '', False),
    ('Office/Utilities', '', False),
    ('Other Expenses', '', False),
    ('Total OpEx', '=SUM(B20:B26)', False),
    ('', '', False),
    ('NET PROFIT', '=B16-B28', True),
    ('Net Margin %', '=IF(B8=0,"",B30/B8)', False),
]
for i, (label, formula, is_header) in enumerate(pl_items, 3):
    cell_a = ws_pl.cell(row=i, column=1, value=label)
    if is_header:
        cell_a.font = Font(bold=True, size=12, color='1A1A2E')
        cell_a.fill = GRAY_FILL
    cell_b = ws_pl.cell(row=i, column=2, value=formula if formula else '')
    if formula and isinstance(formula, str) and formula.startswith('='):
        cell_b.number_format = '0.0%' if '%' in label else '$#,##0.00'
        cell_b.fill = LIGHT_GREEN
    elif not formula and not is_header and label:
        cell_b.fill = INPUT_FILL  # Manual entry cells

# ============================================
# SHEET 7: TARGETS
# ============================================
ws_targets = wb.create_sheet('Targets')
for letter, w in [('A',30),('B',15),('C',15),('D',15),('E',20)]:
    ws_targets.column_dimensions[letter].width = w

ws_targets.cell(row=1, column=1, value='Cresca — MRR Milestones & Targets').font = TITLE_FONT

target_headers = ['Milestone', 'MRR Target', 'Status']
for j, h in enumerate(target_headers, 1):
    ws_targets.cell(row=3, column=j, value=h)
style_header_row(ws_targets, 3, 3)

targets = [
    ['Milestone 1', 2000, '=IF(\'MRR Summary\'!B8>=B4,"✓ Achieved","In Progress")'],
    ['Milestone 2', 5000, '=IF(\'MRR Summary\'!B8>=B5,"✓ Achieved","In Progress")'],
    ['Milestone 3', 10000, '=IF(\'MRR Summary\'!B8>=B6,"✓ Achieved","In Progress")'],
]
for i, row in enumerate(targets, 4):
    for j, val in enumerate(row, 1):
        cell = ws_targets.cell(row=i, column=j, value=val)
        cell.border = THIN_BORDER
        if j == 2:
            cell.number_format = '$#,##0'

# Monthly targets
ws_targets.cell(row=9, column=1, value='Monthly Growth Targets').font = Font(bold=True, size=12)
month_headers = ['Month', 'Target MRR', 'Target Clients', 'New Clients', 'Actual MRR']
for j, h in enumerate(month_headers, 1):
    ws_targets.cell(row=10, column=j, value=h)
style_header_row(ws_targets, 10, 5)

for m in range(1, 13):
    r = 10 + m
    ws_targets.cell(row=r, column=1, value=f'Month {m}').border = THIN_BORDER
    ws_targets.cell(row=r, column=2, value=m * 300).number_format = '$#,##0'
    ws_targets.cell(row=r, column=2).border = THIN_BORDER
    ws_targets.cell(row=r, column=3, value=m * 3).border = THIN_BORDER
    ws_targets.cell(row=r, column=4).fill = INPUT_FILL
    ws_targets.cell(row=r, column=4).border = THIN_BORDER
    ws_targets.cell(row=r, column=5).fill = INPUT_FILL
    ws_targets.cell(row=r, column=5).number_format = '$#,##0'
    ws_targets.cell(row=r, column=5).border = THIN_BORDER

# ============================================
# SHEET 8: CAPACITY PLANNING
# ============================================
ws_cap = wb.create_sheet('Capacity Planning')
ws_cap.column_dimensions['A'].width = 30
ws_cap.column_dimensions['B'].width = 15
ws_cap.column_dimensions['C'].width = 20

ws_cap.cell(row=1, column=1, value='Cresca — Capacity Planning').font = TITLE_FONT

ws_cap.cell(row=3, column=1, value='Current Capacity').font = Font(bold=True, size=12)
cap_items = [
    ('Max Concurrent GROWTH/PRO Builds', 5, ''),
    ('Current Active Builds', '=COUNTIF(Clients!D2:D101,"Onboarding")', ''),
    ('Available Capacity', '=B4-B5', '=IF(B6<=0,"⚠️ AT CAPACITY","✓ Available")'),
    ('Utilization %', '=IF(B4=0,"",B5/B4)', ''),
]
for i, (label, val, note) in enumerate(cap_items, 4):
    ws_cap.cell(row=i, column=1, value=label)
    cell = ws_cap.cell(row=i, column=2, value=val)
    if isinstance(val, str) and val.startswith('='):
        cell.fill = LIGHT_GREEN
    if '%' in label:
        cell.number_format = '0.0%'
    if note:
        ws_cap.cell(row=i, column=3, value=note).fill = LIGHT_GREEN

ws_cap.cell(row=9, column=1, value='Capacity by Package Type').font = Font(bold=True, size=12)
rules = [
    'STARTER: High volume, minimal build time (1-5 days)',
    'GROWTH: 3-5 simultaneous, 1-3 weeks each',
    'PRO AUTOMATION: 1-2 simultaneous, 4-8 weeks each',
    'If Active Builds = 5 → STOP accepting new GROWTH/PRO contracts',
    'If Active Builds = 4 → Accept 1 more max, prioritize high-value',
]
for i, rule in enumerate(rules, 10):
    ws_cap.cell(row=i, column=1, value=rule)

# ============================================
# SHEET 9: MONTHLY HISTORY
# ============================================
ws_hist = wb.create_sheet('Monthly History')
hist_headers = ['Month', 'MRR ($)', 'Active Clients', 'New Clients', 'Cancelled',
                'Churn Rate', 'Setup Revenue ($)', 'Total Commissions ($)',
                'Rep Base Salary ($)', 'Gross Profit ($)', 'Notes']
for j, h in enumerate(hist_headers, 1):
    ws_hist.cell(row=1, column=j, value=h)
    ws_hist.column_dimensions[get_column_letter(j)].width = max(len(h) + 2, 14)
style_header_row(ws_hist, 1, len(hist_headers))

months = ['Jan 2026', 'Feb 2026', 'Mar 2026', 'Apr 2026', 'May 2026', 'Jun 2026',
          'Jul 2026', 'Aug 2026', 'Sep 2026', 'Oct 2026', 'Nov 2026', 'Dec 2026']
for i, month in enumerate(months, 2):
    ws_hist.cell(row=i, column=1, value=month).border = THIN_BORDER
    for c in range(2, 12):
        ws_hist.cell(row=i, column=c).fill = INPUT_FILL
        ws_hist.cell(row=i, column=c).border = THIN_BORDER
        if c in [2, 7, 8, 9, 10]:
            ws_hist.cell(row=i, column=c).number_format = '$#,##0'
        elif c == 6:
            ws_hist.cell(row=i, column=c).number_format = '0.0%'

# ============================================
# SAVE
# ============================================
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Sheets: {wb.sheetnames}")
print("Done!")
