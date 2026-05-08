from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.styles import Font, PatternFill, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = load_workbook('Todo_Directo_Founder_Financial_Dashboard.xlsx')

# ============================================================
# 1. DATA VALIDATION on Clients sheet (columns H, K, Q)
# ============================================================
ws_clients = wb['Clients']

# Remove existing validations if any
ws_clients.data_validations.dataValidation = []

# Status (column H) — rows 3 to 202
dv_status = DataValidation(
    type="list",
    formula1='"Active,Paused,Cancelled"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Status",
    error="Please select: Active, Paused, or Cancelled"
)
dv_status.sqref = 'H3:H202'
ws_clients.add_data_validation(dv_status)

# Billing (column K)
dv_billing = DataValidation(
    type="list",
    formula1='"Monthly,Annual"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Billing",
    error="Please select: Monthly or Annual"
)
dv_billing.sqref = 'K3:K202'
ws_clients.add_data_validation(dv_billing)

# Payment Method (column Q)
dv_payment = DataValidation(
    type="list",
    formula1='"Cash,Transfer,Card,Stripe,Chivo Wallet"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Payment Method",
    error="Please select: Cash, Transfer, Card, Stripe, or Chivo Wallet"
)
dv_payment.sqref = 'Q3:Q202'
ws_clients.add_data_validation(dv_payment)

# ============================================================
# 2. CONDITIONAL FORMATTING
# ============================================================

# --- MRR Summary: Net MRR Change (B9) — red if negative, green if positive ---
ws_mrr = wb['MRR Summary']
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
green_font = Font(color='006100')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
red_font = Font(color='9C0006')

ws_mrr.conditional_formatting.add('B9',
    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill, font=green_font))
ws_mrr.conditional_formatting.add('B9',
    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font))

# --- MRR Summary: Logo Churn Rate (B22) — red if > 3%, green if <= 3% ---
ws_mrr.conditional_formatting.add('B22',
    CellIsRule(operator='greaterThan', formula=['0.03'], fill=red_fill, font=red_font))
ws_mrr.conditional_formatting.add('B22',
    CellIsRule(operator='lessThanOrEqual', formula=['0.03'], fill=green_fill, font=green_font))

# --- P&L: Pre-tax Profit (B19) and Net Profit (B21) — red/green ---
ws_pl = wb['P&L']
for cell_ref in ['B19', 'B21']:
    ws_pl.conditional_formatting.add(cell_ref,
        CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill, font=green_font))
    ws_pl.conditional_formatting.add(cell_ref,
        CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font))

# --- P&L: Gross Margin % (B7) / Net Margin % (B22) — gradient ---
# green if >= 50%, yellow if 25-50%, red if < 25%
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
yellow_font = Font(color='9C6500')

for cell_ref in ['B7', 'B22']:
    ws_pl.conditional_formatting.add(cell_ref,
        CellIsRule(operator='greaterThanOrEqual', formula=['0.5'], fill=green_fill, font=green_font))
    ws_pl.conditional_formatting.add(cell_ref,
        CellIsRule(operator='between', formula=['0.25', '0.4999'], fill=yellow_fill, font=yellow_font))
    ws_pl.conditional_formatting.add(cell_ref,
        CellIsRule(operator='lessThan', formula=['0.25'], fill=red_fill, font=red_font))

# --- Clients: Revenue column (L) — data bars for visual comparison ---
ws_clients.conditional_formatting.add('L3:L202',
    DataBarRule(start_type='min', end_type='max', color='5B9BD5'))

# --- Targets: clients required vs active — highlight if met ---
ws_targets = wb['Targets']
# B13 = clients required. Color green if MRR Summary B4 (active) >= B13
ws_targets.conditional_formatting.add('B13',
    FormulaRule(formula=["'MRR Summary'!B4>=B13"], fill=green_fill, font=green_font))
ws_targets.conditional_formatting.add('B13',
    FormulaRule(formula=["'MRR Summary'!B4<B13"], fill=yellow_fill, font=yellow_font))

# ============================================================
# 3. SHEET PROTECTION (lock formula cells, allow input cells)
# ============================================================

# For MRR Summary — protect all (it's all formulas)
ws_mrr.protection.sheet = True
ws_mrr.protection.password = 'td2026'
ws_mrr.protection.enable()

# For P&L — protect all formulas
ws_pl.protection.sheet = True
ws_pl.protection.password = 'td2026'
ws_pl.protection.enable()

# For Commission Liability — protect formulas, allow rep names (A8:A17)
ws_comm = wb['Commission Liability']
ws_comm.protection.sheet = True
ws_comm.protection.password = 'td2026'
ws_comm.protection.enable()
# Unlock rep name cells
from openpyxl.styles import Protection
unlocked = Protection(locked=False)
for row in range(8, 18):
    cell = ws_comm.cell(row=row, column=1)
    cell.protection = unlocked

# For Targets — protect formulas, allow blue input cells (B4, B7)
ws_targets.protection.sheet = True
ws_targets.protection.password = 'td2026'
ws_targets.protection.enable()
for cell_ref in ['B4', 'B7']:
    ws_targets[cell_ref].protection = unlocked

# For Clients — allow all editing (it's the main input sheet)
# No protection needed

# For Assumptions — allow editing of value column (B3:B12)
ws_assumptions = wb['Assumptions']
ws_assumptions.protection.sheet = True
ws_assumptions.protection.password = 'td2026'
ws_assumptions.protection.enable()
for row in range(3, 13):
    ws_assumptions.cell(row=row, column=2).protection = unlocked

# Update README to note protections
ws_readme = wb['README']
ws_readme['A15'] = "• Formula sheets (MRR Summary, Commission Liability, P&L, Targets) are protected (password: td2026) to prevent accidental formula overwrites. Input cells are unlocked."

# Save
OUTPUT = 'Todo_Directo_Founder_Financial_Dashboard.xlsx'
wb.save(OUTPUT)
print(f'Saved to {OUTPUT}')
print('Changes: data validation (3 cols), conditional formatting (8 rules), sheet protection (5 sheets)')
