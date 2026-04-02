import openpyxl

wb = openpyxl.load_workbook('Todo_Directo_Founder_Financial_Dashboard.xlsx')
print('Sheets:', wb.sheetnames)

# 1: Tax reserve format
ws = wb['Assumptions']
print(f'\n1. B9 format: {ws["B9"].number_format} (expect 0.0%)')

# 2: P&L B14
ws_pl = wb['P&L']
print(f'2. P&L B14: {ws_pl["B14"].value}')

# 3: Data validations
ws_c = wb['Clients']
print(f'3. Validations: {len(ws_c.data_validations.dataValidation)}')
for dv in ws_c.data_validations.dataValidation:
    print(f'   {dv.sqref} -> {dv.formula1}')

# 4: Paused rows
ws_mrr = wb['MRR Summary']
print(f'4. A10={ws_mrr["A10"].value}, A11={ws_mrr["A11"].value}')

# 5: P&L structure
print('5. P&L:')
for r in range(1, ws_pl.max_row+1):
    a = ws_pl.cell(row=r, column=1).value
    b = ws_pl.cell(row=r, column=2).value
    if a or b:
        print(f'   R{r}: {a} | {b}')

# 6: Churn
print('6. MRR tail:')
for r in range(ws_mrr.max_row-4, ws_mrr.max_row+1):
    a = ws_mrr.cell(row=r, column=1).value
    b = ws_mrr.cell(row=r, column=2).value
    if a or b:
        print(f'   R{r}: {a} | {b}')

# 7: Effective MRR
print(f'7. S2={ws_c["S2"].value}, S3={ws_c["S3"].value}')

# 8: Cond fmt
print(f'8. MRR cf={len(ws_mrr.conditional_formatting._cf_rules)}, P&L cf={len(ws_pl.conditional_formatting._cf_rules)}')

# 9: History
ws_hist = wb['Monthly History']
h = [ws_hist.cell(row=3, column=c).value for c in range(1, 13)]
print(f'9. History headers: {h}')

# 10: Protection
for n in ['MRR Summary', 'Commission Liability', 'P&L']:
    print(f'10. {n} protected: {wb[n].protection.sheet}')

# README
print('README:')
ws_r = wb['README']
for r in range(13, 16):
    print(f'   A{r}: {ws_r.cell(row=r, column=1).value}')
