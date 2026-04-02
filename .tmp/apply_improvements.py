import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from copy import copy

wb = openpyxl.load_workbook('Todo_Directo_Founder_Financial_Dashboard.xlsx')

# ============================================================
# Style helpers
# ============================================================
header_fill = PatternFill('solid', start_color='001F4E79')
light_fill = PatternFill('solid', start_color='00D9E1F2')
bold_font = Font(bold=True)
bold_white = Font(bold=True, color='FFFFFF')
currency_fmt = '"$"#,##0.00_);[Red]("$"#,##0.00);"-"_);_(@_)'
pct_fmt = '0.0%'
num_fmt = '#,##0'

# ============================================================
# 1. FIX: Tax reserve % formatting (Assumptions B9)
# ============================================================
ws = wb['Assumptions']
ws['B9'].number_format = pct_fmt
print("✓ Fixed Assumptions B9 format to percentage")

# ============================================================
# 2. FIX: Payment processing fee formula (P&L B14)
#    Include fixed per-txn fee * active clients + apply % to total revenue
# ============================================================
ws_pl = wb['P&L']
ws_pl['B14'] = '=B6*Assumptions!B4 + \'MRR Summary\'!B4*Assumptions!B5'
ws_pl['B14'].number_format = currency_fmt
print("✓ Fixed P&L B14 payment processing formula")

# ============================================================
# 3. FIX: Data validation dropdowns on Clients sheet
# ============================================================
ws_clients = wb['Clients']

# Status dropdown (column H)
dv_status = DataValidation(type="list", formula1='"Active,Paused,Cancelled"', allow_blank=True)
dv_status.error = "Please select Active, Paused, or Cancelled"
dv_status.errorTitle = "Invalid Status"
dv_status.prompt = "Select client status"
dv_status.promptTitle = "Status"
ws_clients.add_data_validation(dv_status)
dv_status.add('H3:H500')

# Billing dropdown (column K)
dv_billing = DataValidation(type="list", formula1='"Monthly,Annual"', allow_blank=True)
dv_billing.error = "Please select Monthly or Annual"
dv_billing.errorTitle = "Invalid Billing"
ws_clients.add_data_validation(dv_billing)
dv_billing.add('K3:K500')

# Payment Method dropdown (column Q)
dv_payment = DataValidation(type="list", formula1='"Cash,Transfer,Card,Stripe"', allow_blank=True)
dv_payment.error = "Please select a valid payment method"
dv_payment.errorTitle = "Invalid Payment Method"
ws_clients.add_data_validation(dv_payment)
dv_payment.add('Q3:Q500')

print("✓ Added data validation dropdowns to Clients sheet (Status, Billing, Payment)")

# ============================================================
# 4. ADD: Paused client tracking to MRR Summary
# ============================================================
ws_mrr = wb['MRR Summary']

# Insert rows after row 9 (after Net MRR Change)
# We'll add at row 10 and 11
ws_mrr.insert_rows(10, 2)

ws_mrr['A10'] = 'Paused Clients'
ws_mrr['A10'].font = bold_font
ws_mrr['B10'] = '=COUNTIF(Clients!H3:H500,"Paused")'

ws_mrr['A11'] = 'Paused MRR (potential reactivation)'
ws_mrr['A11'].font = bold_font
ws_mrr['B11'] = '=SUMIF(Clients!H3:H500,"Paused",Clients!L3:L500)'
ws_mrr['B11'].number_format = currency_fmt

print("✓ Added Paused client tracking to MRR Summary")

# ============================================================
# 5. ADD: Margin % rows to P&L
# ============================================================
# Current layout after our changes:
# Row 6: Total Revenue
# Row 8: Costs header
# Row 9-14: cost lines
# Row 15: Total Costs
# Row 17: Profit header
# Row 18: Pre-tax Profit
# ...

# Add Gross Margin % after Total Revenue (insert row 7)
ws_pl.insert_rows(7, 1)
ws_pl['A7'] = 'Gross Margin %'
ws_pl['A7'].font = Font(bold=True, italic=True)
ws_pl['B7'] = '=IFERROR((B6-B10)/B6,0)'  # (Total Revenue - Client variable costs) / Total Revenue; B10 is client variable costs after the insert
ws_pl['B7'].number_format = pct_fmt

# Now rows shifted: Total Costs is row 16, Pre-tax is row 19, Net Profit is row 21
# Add Net Margin % after Net Profit
# Find current last row
max_row = ws_pl.max_row  # should be 23 after insert
ws_pl.insert_rows(max_row - 1, 1)  # insert before "Remaining for reinvestment"

# Recalculate positions after both inserts:
# Row 1: Monthly Profit & Owner Take
# Row 3: Revenue header
# Row 4: Total MRR
# Row 5: Installation fees
# Row 6: Total Revenue
# Row 7: Gross Margin % (NEW)
# Row 8: (blank)
# Row 9: Costs header
# Row 10: Client variable costs
# Row 11: Sales residual commissions
# Row 12: Tools/APIs
# Row 13: Payroll
# Row 14: Fixed overhead
# Row 15: Payment processing fees
# Row 16: Total Costs
# Row 17: (blank)
# Row 18: Profit header
# Row 19: Pre-tax Profit
# Row 20: Tax reserve
# Row 21: Net Profit
# Row 22: Net Margin % (NEW - inserted)
# Row 23: Owner Take-home
# Row 24: Remaining for reinvestment

ws_pl['A22'] = 'Net Margin %'
ws_pl['A22'].font = Font(bold=True, italic=True)
ws_pl['B22'] = '=IFERROR(B21/B6,0)'
ws_pl['B22'].number_format = pct_fmt

print("✓ Added Gross Margin % and Net Margin % to P&L")

# ============================================================
# 6. ADD: Actual churn metrics to MRR Summary
# ============================================================
# Add after the existing KPIs section (rows shifted due to earlier inserts)
# MRR Summary now has Paused rows at 10-11, so product tag section shifted
# Let's add churn metrics at the end

max_mrr = ws_mrr.max_row + 2
ws_mrr.cell(row=max_mrr, column=1, value='Churn Metrics (actual)').font = bold_font

ws_mrr.cell(row=max_mrr+1, column=1, value='Logo Churn Rate (this month)').font = bold_font
ws_mrr.cell(row=max_mrr+1, column=2).value = '=IFERROR(B8/(B4+B8),0)'
ws_mrr.cell(row=max_mrr+1, column=2).number_format = pct_fmt

ws_mrr.cell(row=max_mrr+2, column=1, value='MRR Lost to Churn ($)').font = bold_font
ws_mrr.cell(row=max_mrr+2, column=2).value = '=SUMIFS(Clients!L3:L500,Clients!H3:H500,"Cancelled",Clients!J3:J500,">="&EOMONTH($E$4,-1)+1,Clients!J3:J500,"<="&EOMONTH($E$4,0))'
ws_mrr.cell(row=max_mrr+2, column=2).number_format = currency_fmt

ws_mrr.cell(row=max_mrr+3, column=1, value='MRR Churn Rate (this month)').font = bold_font
f_churn_mrr = max_mrr + 2
ws_mrr.cell(row=max_mrr+3, column=2).value = f'=IFERROR(B{f_churn_mrr}/(B5+B{f_churn_mrr}),0)'
ws_mrr.cell(row=max_mrr+3, column=2).number_format = pct_fmt

print("✓ Added actual churn metrics to MRR Summary")

# ============================================================
# 7. ADD: Annual billing helper column in Clients
# ============================================================
# Add column S: "Effective MRR" that adjusts for annual billing
ws_clients['S2'] = 'Effective MRR ($)'
ws_clients['S2'].font = bold_white
ws_clients['S2'].fill = PatternFill('solid', start_color='001F4E79')

for r in range(3, 202):
    ws_clients.cell(row=r, column=19).value = f'=IF(K{r}="Annual",L{r}/12,L{r})'
    ws_clients.cell(row=r, column=19).number_format = currency_fmt

print("✓ Added Effective MRR helper column (S) to Clients sheet for annual billing")

# ============================================================
# 8. ADD: Conditional formatting
# ============================================================
# Green fill for positive Net MRR Change, Red for negative
green_fill = PatternFill('solid', start_color='00C6EFCE')
red_fill = PatternFill('solid', start_color='00FFC7CE')
green_font = Font(color='006100')
red_font = Font(color='9C0006')

ws_mrr.conditional_formatting.add('B9',
    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill, font=green_font))
ws_mrr.conditional_formatting.add('B9',
    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font))

# P&L: Pre-tax profit, Net Profit conditional formatting
ws_pl.conditional_formatting.add('B19',
    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill, font=green_font))
ws_pl.conditional_formatting.add('B19',
    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font))
ws_pl.conditional_formatting.add('B21',
    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill, font=green_font))
ws_pl.conditional_formatting.add('B21',
    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font))

print("✓ Added conditional formatting (green/red) on key financial metrics")

# ============================================================
# 9. ADD: Monthly History sheet
# ============================================================
ws_hist = wb.create_sheet('Monthly History')
ws_hist['A1'] = 'Monthly KPI History (paste values each month)'
ws_hist['A1'].font = bold_font

headers_hist = ['Month', 'Active Clients', 'MRR ($)', 'Net Profit ($)', 'Owner Take ($)',
                'New Clients', 'Cancelled', 'Paused', 'Logo Churn %', 'Gross Margin %', 'Net Margin %', 'ARPU ($)']
for i, h in enumerate(headers_hist, 1):
    cell = ws_hist.cell(row=3, column=i, value=h)
    cell.font = bold_white
    cell.fill = PatternFill('solid', start_color='001F4E79')
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Set column widths
col_widths = [14, 14, 14, 14, 14, 12, 12, 10, 12, 14, 12, 12]
for i, w in enumerate(col_widths):
    ws_hist.column_dimensions[chr(65+i)].width = w

print("✓ Created Monthly History tracking sheet")

# ============================================================
# 10. ADD: Sheet protection (formula cells only)
# ============================================================
# Protect MRR Summary, Commission Liability, P&L (keep Assumptions and Clients unlocked for input)
for sheet_name in ['MRR Summary', 'Commission Liability', 'P&L']:
    ws = wb[sheet_name]
    ws.protection.sheet = True
    ws.protection.password = ''  # No password, just UI-level protection
    ws.protection.enable()
    # Unlock all cells first, then lock formula cells
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = openpyxl.styles.Protection(locked=False)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                cell.protection = openpyxl.styles.Protection(locked=True)

print("✓ Added sheet protection on formula sheets")

# ============================================================
# Column width adjustments for readability
# ============================================================
ws_mrr = wb['MRR Summary']
ws_mrr.column_dimensions['A'].width = 38
ws_mrr.column_dimensions['B'].width = 18
ws_mrr.column_dimensions['C'].width = 14
ws_mrr.column_dimensions['D'].width = 16
ws_mrr.column_dimensions['E'].width = 16

ws_pl = wb['P&L']
ws_pl.column_dimensions['A'].width = 38
ws_pl.column_dimensions['B'].width = 18

ws_assumptions = wb['Assumptions']
ws_assumptions.column_dimensions['A'].width = 38
ws_assumptions.column_dimensions['B'].width = 16
ws_assumptions.column_dimensions['C'].width = 55

# ============================================================
# Update README with new instructions
# ============================================================
ws_readme = wb['README']
ws_readme['A13'] = '• Column S in Clients computes Effective MRR (adjusts Annual billing to monthly).'
ws_readme['A14'] = "• The 'Monthly History' sheet should be updated each month by pasting KPI values."
ws_readme['A15'] = '• Formula sheets (MRR Summary, Commission Liability, P&L) are protected to prevent accidental edits.'

# ============================================================
# Save
# ============================================================
wb.save('Todo_Directo_Founder_Financial_Dashboard.xlsx')
print("\n✅ All improvements saved to Todo_Directo_Founder_Financial_Dashboard.xlsx")
