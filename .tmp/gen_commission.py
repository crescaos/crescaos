"""Generate updated Cresca Commission Calculator — 3-Tier Model with 20%/6% rates."""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r'c:\Users\12132\Desktop\DigiFacil\DigiFacil\CRESCA_Commission_Calculator_v2.xlsx'
wb = Workbook()

# Styles
BLUE_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
INPUT_FILL = PatternFill(start_color='EBF5FF', end_color='EBF5FF', fill_type='solid')
GREEN_FILL = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
GRAY_FILL = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
DARK_FILL = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
WHITE_FONT = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color='1A1A2E')
BOLD = Font(bold=True, size=11)
SMALL = Font(size=9, italic=True, color='666666')
MONEY_FMT = '$#,##0.00'
PCT_FMT = '0%'
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

def hdr(ws, row, cols, fill=BLUE_FILL):
    for j, txt in enumerate(cols, 1):
        c = ws.cell(row=row, column=j, value=txt)
        c.fill = fill; c.font = WHITE_FONT; c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = BORDER

def cell(ws, r, c, val, fmt=None, fill=None, font=None, bold=False):
    cl = ws.cell(row=r, column=c, value=val)
    cl.border = BORDER
    if fmt: cl.number_format = fmt
    if fill: cl.fill = fill
    if font: cl.font = font
    if bold: cl.font = Font(bold=True, size=11)
    return cl

# ============================================
# SHEET 1: COTIZADOR DE CLIENTES
# ============================================
ws1 = wb.active
ws1.title = 'Cotizador de Clientes'
for letter, w in [('A',35),('B',20),('C',18),('D',18),('E',20),('F',18),('G',20),('H',18)]:
    ws1.column_dimensions[letter].width = w

ws1.cell(row=1, column=1, value='CRESCA — Cotizador de Clientes (3-Tier Model)').font = TITLE_FONT
ws1.cell(row=2, column=1, value='Sistemas de Crecimiento Empresarial').font = Font(italic=True, size=11)
ws1.cell(row=3, column=1, value='Selecciona los servicios para generar la cotización automática.').font = SMALL

# Package table
hdr(ws1, 5, ['Paquete / Tier', 'Setup Fee ($)', 'Cuota Mensual ($)', 'Tipo',
             '¿Incluir?', 'Comisión Setup ($)', 'Comisión Residual ($/mes)', 'Ingreso Total 12m ($)'])

packages = [
    ['STARTER — DIY Lite', 0, 9.99, 'Starter'],
    ['STARTER — Starter Assist', 25, 19.99, 'Starter'],
    ['STARTER — Starter Presence', 75, 29.99, 'Starter'],
    ['STARTER — Starter Pro', 150, 49.99, 'Starter'],
    ['GROWTH — Foundation', 250, 50, 'Growth'],
    ['GROWTH — Foundation +1 WF', 350, 75, 'Growth'],
    ['GROWTH — Foundation +2 WF', 450, 100, 'Growth'],
    ['PRO AUTOMATION (promedio)', 1500, 500, 'Pro'],
    ['Add-On: Marketing Digital', 200, 250, 'Add-On'],
    ['Add-On: Workflows (+2)', 0, 25, 'Add-On'],
    ['Add-On: Workflows (+5)', 0, 60, 'Add-On'],
    ['Add-On: Workflows (+8)', 0, 90, 'Add-On'],
]

for i, (name, setup, monthly, ptype) in enumerate(packages):
    r = 6 + i
    cell(ws1, r, 1, name)
    cell(ws1, r, 2, setup, MONEY_FMT, INPUT_FILL)
    cell(ws1, r, 3, monthly, MONEY_FMT, INPUT_FILL)
    cell(ws1, r, 4, ptype)
    cell(ws1, r, 5, 'No', fill=INPUT_FILL)
    # Commission Setup = Setup * 20% if included
    cell(ws1, r, 6, f'=IF(UPPER(E{r})="SI",B{r}*0.20,IF(UPPER(E{r})="SÍ",B{r}*0.20,0))', MONEY_FMT, GREEN_FILL)
    # Residual = Monthly * 6% if included
    cell(ws1, r, 7, f'=IF(UPPER(E{r})="SI",C{r}*0.06,IF(UPPER(E{r})="SÍ",C{r}*0.06,0))', MONEY_FMT, GREEN_FILL)
    # Total 12-month = Setup commission + (Residual * 12)
    cell(ws1, r, 8, f'=IF(UPPER(E{r})="SI",F{r}+G{r}*12,IF(UPPER(E{r})="SÍ",F{r}+G{r}*12,0))', MONEY_FMT, GREEN_FILL)

# Drop-down for include
dv_include = DataValidation(type='list', formula1='"Sí,No"')
ws1.add_data_validation(dv_include)
last_pkg_row = 6 + len(packages) - 1
for r in range(6, last_pkg_row + 1):
    dv_include.add(ws1.cell(row=r, column=5))

# Totals
r_total = last_pkg_row + 2
ws1.cell(row=r_total, column=1, value='📊 TOTALES DE LA COTIZACIÓN').font = BOLD

cell(ws1, r_total, 2, f'=SUMPRODUCT((UPPER(E6:E{last_pkg_row})="SÍ")*B6:B{last_pkg_row})+SUMPRODUCT((UPPER(E6:E{last_pkg_row})="SI")*B6:B{last_pkg_row})', MONEY_FMT, YELLOW_FILL, bold=True)
cell(ws1, r_total, 3, f'=SUMPRODUCT((UPPER(E6:E{last_pkg_row})="SÍ")*C6:C{last_pkg_row})+SUMPRODUCT((UPPER(E6:E{last_pkg_row})="SI")*C6:C{last_pkg_row})', MONEY_FMT, YELLOW_FILL, bold=True)
cell(ws1, r_total, 6, f'=SUM(F6:F{last_pkg_row})', MONEY_FMT, YELLOW_FILL, bold=True)
cell(ws1, r_total, 7, f'=SUM(G6:G{last_pkg_row})', MONEY_FMT, YELLOW_FILL, bold=True)
cell(ws1, r_total, 8, f'=SUM(H6:H{last_pkg_row})', MONEY_FMT, YELLOW_FILL, bold=True)

# Client summary section
r = r_total + 2
ws1.cell(row=r, column=1, value='📝 RESUMEN PARA EL CLIENTE').font = BOLD
cell(ws1, r+1, 1, 'Inversión Inicial (Setup Fee):')
cell(ws1, r+1, 3, f'=B{r_total}', MONEY_FMT, GREEN_FILL)
cell(ws1, r+2, 1, 'Cuota Mensual Total:')
cell(ws1, r+2, 3, f'=C{r_total}', MONEY_FMT, GREEN_FILL)

r = r_total + 6
ws1.cell(row=r, column=1, value='💰 TUS GANANCIAS (Rep Fundador ES)').font = BOLD
cell(ws1, r+1, 1, 'Salario Base:')
cell(ws1, r+1, 3, 600, MONEY_FMT, GREEN_FILL)
cell(ws1, r+2, 1, 'Comisión de Setup (20%):')
cell(ws1, r+2, 3, f'=F{r_total}', MONEY_FMT, GREEN_FILL)
cell(ws1, r+3, 1, 'Comisión Residual Mensual (6%):')
cell(ws1, r+3, 3, f'=G{r_total}', MONEY_FMT, GREEN_FILL)
cell(ws1, r+4, 1, 'Total Mes 1 (Base + Setup + Residual):')
cell(ws1, r+4, 3, f'=600+F{r_total}+G{r_total}', MONEY_FMT, GREEN_FILL)

r = r_total + 12
ws1.cell(row=r, column=1, value='⚠️ REGLAS DE COMPENSACIÓN').font = BOLD
cell(ws1, r+1, 1, '• Salario base: $600/mes (garantizado)')
cell(ws1, r+2, 1, '• Comisión Setup: 20% del Setup Fee cobrado')
cell(ws1, r+3, 1, '• Residual: 6% de la cuota mensual (cada mes que permanezca activo)')
cell(ws1, r+4, 1, '• Clawback: Si cancela en 60 días por mal fit → 50% de comisión setup devuelta')
cell(ws1, r+5, 1, '• Comisiones solo sobre dinero COBRADO (no facturado)')
cell(ws1, r+6, 1, '• No hay comisión sobre ad spend, fees de terceros ni reembolsos')

# ============================================
# SHEET 2: CALCULADORA DE COMISIONES
# ============================================
ws2 = wb.create_sheet('Calculadora de Comisiones')
for letter, w in [('A',6),('B',22),('C',22),('D',18),('E',18),('F',15),('G',18),('H',18),('I',20)]:
    ws2.column_dimensions[letter].width = w

ws2.cell(row=1, column=1, value='CRESCA — CALCULADORA DE COMISIONES (Rep Fundador ES)').font = TITLE_FONT
ws2.merge_cells('A1:I1')
ws2.cell(row=2, column=1, value='Ingresa cada venta cerrada. Las comisiones se calculan automáticamente con 20%/6%.').font = SMALL

# Rep info
ws2.cell(row=4, column=1, value='Rep:').font = BOLD
ws2.cell(row=4, column=2).fill = INPUT_FILL
ws2.cell(row=4, column=4, value='Base:').font = BOLD
ws2.cell(row=4, column=5, value=600).fill = INPUT_FILL
ws2.cell(row=4, column=5).number_format = MONEY_FMT
ws2.cell(row=4, column=7, value='Mes:').font = BOLD
ws2.cell(row=4, column=8).fill = INPUT_FILL

# Client entries header
hdr(ws2, 6, ['#', 'Cliente', 'Paquete', 'Setup Fee ($)', 'Cuota Mensual ($)',
             '¿Cerrado?', 'Comisión Setup ($)', 'Residual ($/mes)', 'Total 12m ($)'])

# Package validation
dv_pkg = DataValidation(type='list', formula1='"STARTER DIY Lite,STARTER Assist,STARTER Presence,STARTER Pro,GROWTH Foundation,GROWTH +1 WF,GROWTH +2 WF,PRO AUTOMATION,Add-On Marketing,Add-On Workflows"')
dv_closed = DataValidation(type='list', formula1='"Sí,No"')
ws2.add_data_validation(dv_pkg)
ws2.add_data_validation(dv_closed)

# 30 entry rows
for i in range(30):
    r = 7 + i
    cell(ws2, r, 1, i + 1)
    cell(ws2, r, 2, '', fill=INPUT_FILL)
    cell(ws2, r, 3, '', fill=INPUT_FILL)
    dv_pkg.add(ws2.cell(row=r, column=3))
    cell(ws2, r, 4, '', MONEY_FMT, INPUT_FILL)
    cell(ws2, r, 5, '', MONEY_FMT, INPUT_FILL)
    cell(ws2, r, 6, '', fill=INPUT_FILL)
    dv_closed.add(ws2.cell(row=r, column=6))
    # Commission Setup = Setup * 20% if closed
    cell(ws2, r, 7, f'=IF(UPPER(F{r})="SÍ",D{r}*0.20,IF(UPPER(F{r})="SI",D{r}*0.20,0))', MONEY_FMT, GREEN_FILL)
    # Residual = Monthly * 6% if closed
    cell(ws2, r, 8, f'=IF(UPPER(F{r})="SÍ",E{r}*0.06,IF(UPPER(F{r})="SI",E{r}*0.06,0))', MONEY_FMT, GREEN_FILL)
    # Total 12-month projection
    cell(ws2, r, 9, f'=IF(UPPER(F{r})="SÍ",G{r}+H{r}*12,IF(UPPER(F{r})="SI",G{r}+H{r}*12,0))', MONEY_FMT, GREEN_FILL)

# Summary section
r_sum = 38
ws2.cell(row=r_sum, column=1, value='📊 RESUMEN DEL MES').font = BOLD
ws2.merge_cells(f'A{r_sum}:C{r_sum}')

cell(ws2, r_sum+1, 2, 'Total Clientes Cerrados:', font=BOLD)
cell(ws2, r_sum+1, 5, '=COUNTIF(F7:F36,"Sí")+COUNTIF(F7:F36,"Si")', fill=YELLOW_FILL)

cell(ws2, r_sum+2, 2, 'Total Setup Fees:', font=BOLD)
cell(ws2, r_sum+2, 5, '=SUMPRODUCT((UPPER(F7:F36)="SÍ")*D7:D36)+SUMPRODUCT((UPPER(F7:F36)="SI")*D7:D36)', MONEY_FMT, YELLOW_FILL)

cell(ws2, r_sum+3, 2, 'Total MRR Generado:', font=BOLD)
cell(ws2, r_sum+3, 5, '=SUMPRODUCT((UPPER(F7:F36)="SÍ")*E7:E36)+SUMPRODUCT((UPPER(F7:F36)="SI")*E7:E36)', MONEY_FMT, YELLOW_FILL)

cell(ws2, r_sum+5, 2, '💰 TUS GANANCIAS', font=BOLD)

cell(ws2, r_sum+6, 2, 'Salario Base:')
cell(ws2, r_sum+6, 5, '=E4', MONEY_FMT, GREEN_FILL)

cell(ws2, r_sum+7, 2, 'Total Comisión Setup (20%):')
cell(ws2, r_sum+7, 5, '=SUM(G7:G36)', MONEY_FMT, GREEN_FILL)

cell(ws2, r_sum+8, 2, 'Total Residual Mensual (6%):')
cell(ws2, r_sum+8, 5, '=SUM(H7:H36)', MONEY_FMT, GREEN_FILL)

cell(ws2, r_sum+10, 2, '🏆 TOTAL GANANCIAS DEL MES:', font=Font(bold=True, size=12))
cell(ws2, r_sum+10, 5, f'=E{r_sum+6}+E{r_sum+7}+E{r_sum+8}', MONEY_FMT, GREEN_FILL, font=Font(bold=True, size=12, color='10B981'))

cell(ws2, r_sum+12, 2, 'Proyección 12 Meses (Base + Residual acum.):', font=Font(bold=True, size=11))
cell(ws2, r_sum+12, 5, f'=E{r_sum+6}*12+E{r_sum+7}+E{r_sum+8}*12', MONEY_FMT, GREEN_FILL)

# ============================================
# SHEET 3: RESUMEN DE BONOS (optional tracking)
# ============================================
ws3 = wb.create_sheet('Referencia de Paquetes')
for letter, w in [('A',35),('B',18),('C',18),('D',18),('E',22)]:
    ws3.column_dimensions[letter].width = w

ws3.cell(row=1, column=1, value='CRESCA — Referencia de Paquetes y Precios').font = TITLE_FONT
ws3.merge_cells('A1:E1')

hdr(ws3, 3, ['Paquete / Tier', 'Setup Fee', 'Mensual', 'Workflows', 'Notas'])

ref_data = [
    ['STARTER — DIY Lite', '$0', '$9.99', '1 Lite', 'Ebook + templates'],
    ['STARTER — Starter Assist', '$0–$49', '$19.99', '1 Lite', '+ Google Business'],
    ['STARTER — Starter Presence', '$49–$99', '$29.99', '1 Lite', '+ Link hub/booking lite'],
    ['STARTER — Starter Pro', '$99–$199', '$49.99', '1 Lite', '+ Micro-sitio 1 página'],
    ['GROWTH — Foundation', '$199–$299', '$50', '1 Lite', 'Website + leads + Google'],
    ['GROWTH — Foundation +1 WF', '$299–$399', '$75', '1 Full', '+ 1 Full Workflow'],
    ['GROWTH — Foundation +2 WF', '$399–$499', '$100', '2 Full', '+ 2 Full Workflows'],
    ['PRO AUTOMATION', 'Cotización', 'Cotización', '4+', 'Auditoría + monitoreo'],
    ['Add-On: Marketing Digital', 'Variable', 'Variable', '—', 'Solo con Growth/Pro'],
    ['Add-On: +1 Workflow', '$0', '$15/mes', '+1 Full', 'Growth add-on'],
    ['Add-On: +2 Workflows', '$0', '$25/mes', '+2 Full', 'Bundle discount'],
    ['Add-On: +5 Workflows', '$0', '$60/mes', '+5 Full', 'Bundle discount'],
    ['Add-On: +8 Workflows', '$0', '$90/mes', '+8 Full', 'Bundle discount'],
]

for i, row in enumerate(ref_data, 4):
    for j, val in enumerate(row, 1):
        cell(ws3, i, j, val)

r = 4 + len(ref_data) + 2
ws3.cell(row=r, column=1, value='Compensación Rep Fundador ES').font = Font(bold=True, size=12)
comp_data = [
    ['Salario Base', '$600/mes (garantizado)'],
    ['Comisión Setup', '20% del fee cobrado'],
    ['Comisión Residual', '6% del MRR cobrado'],
    ['Clawback', '50% setup si cancela en 60 días'],
    ['Pago', 'Solo sobre dinero cobrado'],
]
for i, (label, val) in enumerate(comp_data, r+1):
    cell(ws3, i, 1, label)
    cell(ws3, i, 2, val)

# ============================================
# SAVE
# ============================================
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Sheets: {wb.sheetnames}")
print("Done!")
