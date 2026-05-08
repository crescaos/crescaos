"""Read Commission Calculator in full detail."""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

path = r'c:\Users\12132\Desktop\DigiFacil\DigiFacil\Todo_Directo_Commission_Calculator_v2.xlsx'

# Read with formulas too
wb_f = load_workbook(path)
wb_d = load_workbook(path, data_only=True)

for sname in wb_f.sheetnames:
    ws_f = wb_f[sname]
    ws_d = wb_d[sname]
    print(f"\n{'='*70}")
    print(f"Sheet: {sname} ({ws_f.max_row} rows x {ws_f.max_column} cols)")
    print(f"{'='*70}")
    for r in range(1, min(ws_f.max_row+1, 50)):
        line_parts = []
        for c in range(1, min(ws_f.max_column+1, 15)):
            cell_f = ws_f.cell(row=r, column=c)
            cell_d = ws_d.cell(row=r, column=c)
            if cell_f.value is not None:
                formula = str(cell_f.value)[:60]
                data = str(cell_d.value)[:30] if cell_d.value is not None else ''
                coord = cell_f.coordinate
                if formula != data and data:
                    line_parts.append(f"{coord}=[{formula}]={data}")
                else:
                    line_parts.append(f"{coord}={formula}")
        if line_parts:
            print("  ".join(line_parts))
