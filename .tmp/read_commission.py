"""Read Commission Calculator structure."""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

path = r'c:\Users\12132\Desktop\DigiFacil\DigiFacil\Todo_Directo_Commission_Calculator_v2.xlsx'
wb = load_workbook(path, data_only=True)
print(f"Sheets: {wb.sheetnames}\n")

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"{'='*60}")
    print(f"Sheet: {sname} ({ws.max_row} rows x {ws.max_column} cols)")
    print(f"{'='*60}")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 35), values_only=False):
        vals = []
        for c in row:
            v = c.value
            if v is not None:
                vals.append(f"{c.coordinate}={str(v)[:50]}")
        if vals:
            print("  ".join(vals))
    print()
