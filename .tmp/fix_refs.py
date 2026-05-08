import openpyxl

wb = openpyxl.load_workbook('Todo_Directo_Founder_Financial_Dashboard.xlsx')
ws = wb['P&L']

# Print current state
print("BEFORE fixes:")
for r in range(1, ws.max_row+1):
    a = ws.cell(row=r, column=1).value
    b = ws.cell(row=r, column=2).value
    if a or b:
        print(f"  R{r}: {a} | {b}")

# Fix formulas that didn't auto-adjust after row insert
# Current layout:
# R4:  Total MRR           | ='MRR Summary'!B5
# R5:  Installation fees   | =SUMIFS(...)
# R6:  Total Revenue       | =B4+B5  ✓
# R7:  Gross Margin %      | =IFERROR((B6-B10)/B6,0)  ✓ (B10 = client var costs)
# R9:  Costs header
# R10: Client variable costs
# R11: Sales residual commissions
# R12: Tools/APIs
# R13: Payroll
# R14: Fixed overhead
# R15: Payment processing fees
# R16: Total Costs         | =SUM(B9:B14) → WRONG, should be =SUM(B10:B15)
# R18: Profit header
# R19: Pre-tax Profit      | =B6-B15 → WRONG, should be =B6-B16
# R20: Tax reserve         | =MAX(0,B18*Assumptions!B9) → WRONG, should be B19
# R21: Net Profit          | =B18-B19 → WRONG, should be =B19-B20
# R22: Net Margin %        | =IFERROR(B21/B6,0) ✓ 
# R23: Owner Take-home     | =MAX(0,B20*Assumptions!B12) → WRONG, should be B21
# R24: Remaining           | =B20-B21 → WRONG, should be =B21-B23

currency_fmt = '"$"#,##0.00_);[Red]("$"#,##0.00);"-"_);_(@_)'

# Fix Total Costs
ws['B16'] = '=SUM(B10:B15)'
ws['B16'].number_format = currency_fmt
ws['B16'].font = openpyxl.styles.Font(bold=True)

# Fix Pre-tax Profit
ws['B19'] = '=B6-B16'
ws['B19'].number_format = currency_fmt

# Fix Tax reserve
ws['B20'] = '=MAX(0,B19*Assumptions!B9)'
ws['B20'].number_format = currency_fmt

# Fix Net Profit
ws['B21'] = '=B19-B20'
ws['B21'].number_format = currency_fmt
ws['B21'].font = openpyxl.styles.Font(bold=True)

# Fix Owner Take-home
ws['B23'] = '=MAX(0,B21*Assumptions!B12)'
ws['B23'].number_format = currency_fmt
ws['B23'].font = openpyxl.styles.Font(bold=True)

# Fix Remaining for reinvestment
ws['B24'] = '=B21-B23'
ws['B24'].number_format = currency_fmt

wb.save('Todo_Directo_Founder_Financial_Dashboard.xlsx')

# Print after state
wb2 = openpyxl.load_workbook('Todo_Directo_Founder_Financial_Dashboard.xlsx')
ws2 = wb2['P&L']
print("\nAFTER fixes:")
for r in range(1, ws2.max_row+1):
    a = ws2.cell(row=r, column=1).value
    b = ws2.cell(row=r, column=2).value
    if a or b:
        print(f"  R{r}: {a} | {b}")

print("\n✅ All P&L formula references corrected")
